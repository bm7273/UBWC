"""Build kit.db from data/Kit Inventory.xlsx (the current ground-truth sheet).

Idempotent: drops everything and rebuilds from schema.sql, then seeds the
`items` table from each sheet and the `faults` table from the sheet's Faults
column. Run directly (`python migrate.py`) or via db.rebuild_from_xlsx().
"""
import re
import sqlite3
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DB_PATH = Path.home() / "Library" / "Application Support" / "UBWC" / "kit.db"
SCHEMA_PATH = ROOT / "schema.sql"
XLSX_PATH = ROOT / "data" / "Kit Inventory.xlsx"

# Per-sheet import config. `header_map` maps the sheet's column header to an
# `items` column. `component_type` is either a literal or a callable taking the
# row's "Type" cell value (used to split Sails + Wings into sail/wing).
SHEETS = {
    "Boards": {
        "component_type": "board",
        "header_map": {
            "Manufacturer": "manufacturer",
            "Model": "model",
            "Type": "type",
            "Size (L)": "size_l",
            "Condition": "condition",
            "Location": "location",
        },
    },
    "Sails + Wings": {
        "component_type": lambda t: "wing" if str(t).strip().lower() == "wing" else "sail",
        "header_map": {
            "Manufacturer": "manufacturer",
            "Model": "model",
            "Type": "type",
            "Size (m^2)": "size_m2",
            "Condition": "condition",
            "Location": "location",
            # "Mast Length" + "Extension" are summed into luff_cm below (the DB
            # stores the luff, not a specific recommended mast + extension).
            "Boom": "req_boom_cm",
            "Cams": "cams",
        },
    },
    "Booms": {
        "component_type": "boom",
        "header_map": {
            "Manufacturer": "manufacturer",
            "Model": "model",
            "Type": "type",
            "Min size": "min_size_cm",
            "Max size": "max_size_cm",
            "Condition": "condition",
            "Location": "location",
        },
    },
    "Masts": {
        "component_type": "mast",
        "header_map": {
            "Manufacturer": "manufacturer",
            "Model": "model",
            "Type": "type",
            "Size": "length_cm",
            "Condition": "condition",
            "Location": "location",
        },
    },
    # The sheet has no Extensions tab: they are Misc rows whose Type cell reads
    # "Extension". They are a component type of their own in the DB, so they are
    # split out here and their written travel ("0-30cm") parsed into columns.
    "Misc": {
        "component_type": lambda t: (
            "ext" if str(t or "").strip().lower().startswith("extension") else "misc"),
        "header_map": {
            "Manufacturer": "manufacturer",
            "Model": "model",
            "Type": "type",
            "Size": "size_generic",
            "Condition": "condition",
            "Location": "location",
        },
    },
}

# "0-30cm" in a Size cell, and the RDM/SDM the sheet writes into a Model or Type
# cell. Both become columns on import (see db.py "Derived item facts").
RANGE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)")
DIAMETER_RE = re.compile(r"\b(RDM|SDM)\b", re.IGNORECASE)


def _clean(value):
    """Normalise a cell value: strip strings, drop blanks to None."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def rebuild(db_path: Path = DB_PATH, xlsx_path: Path = XLSX_PATH) -> dict:
    """Drop + recreate the DB from schema.sql and seed it from the xlsx.

    Returns a small summary dict of counts for verification.
    """
    if db_path.exists():
        db_path.unlink()

    conn = sqlite3.connect(db_path)
    conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    item_count = 0
    fault_count = 0

    for sheet_name, cfg in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [(_clean(h) if h is not None else None) for h in rows[0]]
        col_index = {h: i for i, h in enumerate(headers) if h is not None}

        for row in rows[1:]:
            # Skip fully-empty rows.
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                continue

            record = {}
            for header, field in cfg["header_map"].items():
                if header in col_index:
                    record[field] = _clean(row[col_index[header]])

            # cams -> 0/1
            if "cams" in record and record["cams"] is not None:
                record["cams"] = 1 if record["cams"] in (True, 1, "TRUE", "True", "yes", "Yes") else 0

            # Sails: the sheet stores "Mast Length" + "Extension"; the DB stores
            # their sum as luff_cm (any mast + extension totalling this fits).
            if "Mast Length" in col_index or "Extension" in col_index:
                ml = _clean(row[col_index["Mast Length"]]) if "Mast Length" in col_index else None
                ex = _clean(row[col_index["Extension"]]) if "Extension" in col_index else None
                if ml is not None or ex is not None:
                    record["luff_cm"] = (ml or 0) + (ex or 0)

            # Resolve component_type (literal or based on the Type cell).
            ctype = cfg["component_type"]
            record["component_type"] = ctype(record.get("type")) if callable(ctype) else ctype

            # Extensions: the written travel becomes a range, and the Type cell
            # has said all it can once component_type is 'ext'.
            if record["component_type"] == "ext":
                match = RANGE_RE.search(str(record.pop("size_generic", "") or ""))
                record["ext_min_cm"] = float(match.group(1)) if match else 0
                record["ext_max_cm"] = float(match.group(2)) if match else None
                record["type"] = None

            # Diameter, wherever the sheet happens to state it. It is a column
            # now, so it is read once here rather than by every caller.
            if record["component_type"] in ("mast", "ext", "sail"):
                haystack = " ".join(str(record.get(f) or "") for f in
                                    ("model", "type", "notes"))
                found = DIAMETER_RE.search(haystack)
                if found:
                    record["diameter"] = found.group(1).upper()

            fields = list(record.keys())
            placeholders = ", ".join("?" for _ in fields)
            cur = conn.execute(
                f"INSERT INTO items ({', '.join(fields)}) VALUES ({placeholders})",
                [record[f] for f in fields],
            )
            item_id = cur.lastrowid
            item_count += 1

            # Seed a fault row from the sheet's Faults column, if present.
            faults_val = _clean(row[col_index["Faults"]]) if "Faults" in col_index else None
            if faults_val:
                conn.execute(
                    "INSERT INTO faults (item_id, description, reported_by) VALUES (?, ?, ?)",
                    (item_id, faults_val, "import"),
                )
                fault_count += 1

    conn.commit()

    summary = {"items": item_count, "faults": fault_count, "by_type": {}}
    for ctype, n in conn.execute(
        "SELECT component_type, COUNT(*) FROM items GROUP BY component_type"
    ):
        summary["by_type"][ctype] = n
    conn.close()
    return summary


if __name__ == "__main__":
    result = rebuild()
    print(f"Built {DB_PATH}")
    print(f"  items:  {result['items']}  {result['by_type']}")
    print(f"  faults: {result['faults']}")
